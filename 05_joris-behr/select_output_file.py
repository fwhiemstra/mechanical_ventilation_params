"""
Output_file selection
- Reads and writes results to existing .xlsx file
- OR creates a new .xlsx file

Author: Sanne van Deelen
Date: February 2021

Modified by Bart Keulen
Date: October 2021

Modified by Anne Meester
Date: February 2022
"""
from pathlib import PurePath
import xlsxwriter
from openpyxl import load_workbook

from constants import PRESSURE_TYPE, OUTPUT_FILE, FS


def select_output_file(output_option, new_output_name, output_xlsx_file, input_filename,
                       patient_id, pressure_type, data_length, rr_,
                       mean_tidal_volume, mech_power, mean_peep,
                       ptp_es_mean, ptp_tp_mean, wob_es_mean, wob_tp_mean, es_loop_power, aw_loop_power,
                       tp_loop_power, tp_peak_mean, tp_swing_mean,
                       standard_deviations, standard_errors):
    """
    Returns output file and name
    """

    # define pressure type selections for documentation
    if pressure_type == PRESSURE_TYPE.AIRWAY:
        pressure_type_txt = 'Airway pressure'
    elif pressure_type == PRESSURE_TYPE.TRANSPULMONARY:
        pressure_type_txt = 'Airway pressure + transpulmonary pressure'

    # calculate segment time
    segment_time = round(data_length / FS / 60)

    if output_option == OUTPUT_FILE.NEW_FILE:
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(new_output_name + '.xlsx')
        worksheet = workbook.add_worksheet()

        # format
        bold = workbook.add_format({'bold': 1})

        # columns
        worksheet.set_column(1, 1, 30)
        worksheet.write('A1', 'Filename', bold)
        worksheet.write('B1', 'Patient ID', bold)
        worksheet.write('C1', 'Pressure Type', bold)
        worksheet.write('D1', 'Segment Length [min]', bold)
        worksheet.write('E1', 'Respiratory Rate [per min]', bold)
        worksheet.write('F1', 'Tidal Volume [mL]', bold)
        worksheet.write('G1', 'PEEP [cmH2O]', bold)
        worksheet.write('H1', 'Airway power [J/min]', bold)
        worksheet.write('I1', 'SD airway power', bold)
        worksheet.write('J1', 'SE airway power', bold)
        worksheet.write('K1', 'Esophageal HA [J/min]', bold)
        worksheet.write('L1', 'SD Esophageal HA', bold)
        worksheet.write('M1', 'SE Esophageal HA', bold)
        worksheet.write('N1', 'Airway HA [J/min]', bold)
        worksheet.write('O1', 'SD Airway HA', bold)
        worksheet.write('P1', 'SE Airway HA', bold)
        worksheet.write('Q1', 'Esophageal PTP [cmH2O.s/min]', bold)
        worksheet.write('R1', 'SD Esophageal PTP', bold)
        worksheet.write('S1', 'SE Esophageal PTP', bold)
        worksheet.write('T1', 'Esophageal POB [J/min]', bold)
        worksheet.write('U1', 'SD Esophageal POB', bold)
        worksheet.write('V1', 'SE Esophageal POB', bold)
        worksheet.write('W1', 'Transpulmonary PTP [cmH2O.s/min]', bold)
        worksheet.write('X1', 'SD Transpulmonary PTP', bold)
        worksheet.write('Y1', 'SE Transpulmonary PTP', bold)
        worksheet.write('Z1', 'Transpulmonary power [J/min]', bold)
        worksheet.write('AA1', 'SD Transpulmonary power', bold)
        worksheet.write('AB1', 'SE Transpulmonary power', bold)
        worksheet.write('AC1', 'Transpulmonary HA [J/min]', bold)
        worksheet.write('AD1', 'SD Transpulmonary HA', bold)
        worksheet.write('AE1', 'SE Transpulmonary HA', bold)
        worksheet.write('AF1', 'Transpulmonary Peak Pressure [cmH2O]', bold)
        worksheet.write('AG1', 'SD Transpulmonary Peak Pressure', bold)
        worksheet.write('AH1', 'SE Transpulmonary Peak Pressure', bold)
        worksheet.write('AI1', 'Transpulmonary Pressure Swing [cmH2O]', bold)
        worksheet.write('AJ1', 'SD Transpulmonary Pressure Swing', bold)
        worksheet.write('AK1', 'SE Transpulmonary Pressure Swing', bold)

        # write data to worksheet
        columns_ = (
            input_filename,
            patient_id,
            pressure_type_txt,
            segment_time,
            rr_,
            mean_tidal_volume,
            mean_peep,
            mech_power,
            standard_deviations[0],
            standard_errors[0],
            es_loop_power,
            standard_deviations[1],
            standard_errors[1],
            aw_loop_power,
            standard_deviations[2],
            standard_errors[2],
            ptp_es_mean,
            standard_deviations[6],
            standard_errors[6],
            wob_es_mean,
            standard_deviations[5],
            standard_errors[5],
            ptp_tp_mean,
            standard_deviations[7],
            standard_errors[7],
            wob_tp_mean,
            standard_deviations[4],
            standard_errors[4],
            tp_loop_power,
            standard_deviations[3],
            standard_errors[3],
            tp_peak_mean,
            standard_deviations[8],
            standard_errors[8],
            tp_swing_mean,
            standard_deviations[9],
            standard_errors[9],
        )

        # define start row and column
        row = 1
        col = 0

        # iterate over the data and write it out column by column
        for item in columns_:
            worksheet.write(row, col, item)
            col += 1

        workbook.close()

        output_file = 'new file created'
        output_filename = new_output_name + '.xlsx'

    elif output_option == OUTPUT_FILE.EXISTING_FILE:
        output_file = output_xlsx_file[0]

        list_output = [input_filename,
            patient_id,
            pressure_type_txt,
            segment_time,
            rr_,
            mean_tidal_volume,
            mean_peep,
            mech_power,
            standard_deviations[0],
            standard_errors[0],
            es_loop_power,
            standard_deviations[1],
            standard_errors[1],
            aw_loop_power,
            standard_deviations[2],
            standard_errors[2],
            ptp_es_mean,
            standard_deviations[6],
            standard_errors[6],
            wob_es_mean,
            standard_deviations[5],
            standard_errors[5],
            ptp_tp_mean,
            standard_deviations[7],
            standard_errors[7],
            wob_tp_mean,
            standard_deviations[4],
            standard_errors[4],
            tp_loop_power,
            standard_deviations[3],
            standard_errors[3],
            tp_peak_mean,
            standard_deviations[8],
            standard_errors[8],
            tp_swing_mean,
            standard_deviations[9],
            standard_errors[9]]

        # re-open and append
        file = load_workbook(output_file)
        ws_ = file.active
        ws_.append(list_output)
        file.save(output_file)
        output_filename = PurePath(output_file).stem + PurePath(output_file).suffix

    return output_file, output_filename
