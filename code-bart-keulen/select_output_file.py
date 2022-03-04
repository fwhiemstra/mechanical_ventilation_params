"""
Output_file selection
- Reads and writes results to existing .xlsx file
- OR creates a new .xlsx file

Author: Sanne van Deelen
Date: February 2021

Modified by Bart Keulen
Date: October 2021
"""
from pathlib import PurePath
import xlsxwriter
from openpyxl import load_workbook

from constants import PRESSURE_TYPE, VENTILATION_MODE, INSP_HOLD, OUTPUT_FILE, FS


def select_output_file(output_option, new_output_name, output_xlsx_file, input_filename,
                       patient_id, ventilation_mode, pressure_type, data_length, rr_,
                       mean_tidal_volume, mech_power, dyn_mech_power, insp_hold, mean_peep,
                       ptp_es_mean, ptp_tp_mean, wob_es_mean, wob_tp_mean, aw_loop_power,
                       tp_loop_power, tp_peak_mean, tp_swing_mean,
                       standard_deviations, standard_errors):
    """
    Returns output file and name
    """
    # define ventilation mode selections for documentation
    if ventilation_mode == VENTILATION_MODE.VCV:
        ventilation_mode_txt = 'VCV - volume controlled ventilation'
    elif ventilation_mode == VENTILATION_MODE.PCV:
        ventilation_mode_txt = 'PCV - pressure controlled ventilation'

    # define pressure type selections for documentation
    if pressure_type == PRESSURE_TYPE.AIRWAY:
        pressure_type_txt = 'Airway pressure'
    elif pressure_type == PRESSURE_TYPE.TRANSPULMONARY:
        pressure_type_txt = 'Airway pressure + transpulmonary pressure'

    # define inspiratory hold selections for documentation
    if insp_hold == INSP_HOLD.ZERO_PERCENT:
        insp_hold_txt = '0%'
    elif insp_hold == INSP_HOLD.TEN_PERCENT:
        insp_hold_txt = '10%'
    elif insp_hold == INSP_HOLD.NONE:
        insp_hold_txt = '-'

    # calculate segment time
    segment_time = round(data_length / FS / 60)

    if output_option == OUTPUT_FILE.NEW_FILE:
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(new_output_name + '.xlsx')
        worksheet = workbook.add_worksheet()

        # format
        bold = workbook.add_format({'bold': 1})

        # columns
        worksheet.set_column(1, 1, 27)
        worksheet.write('A1', 'Filename', bold)
        worksheet.write('B1', 'Patient ID', bold)
        worksheet.write('C1', 'Ventilation Mode', bold)
        worksheet.write('D1', 'Inspiratory Hold', bold)
        worksheet.write('E1', 'Pressure Type', bold)
        worksheet.write('F1', 'Segment Length [min]', bold)
        worksheet.write('G1', 'Respiratory Rate [per min]', bold)
        worksheet.write('H1', 'Tidal Volume [mL]', bold)
        worksheet.write('I1', 'PEEP [cmH2O]', bold)
        worksheet.write('J1', 'MP [J/min]', bold)
        worksheet.write('K1', 'SD MP', bold)
        worksheet.write('L1', 'SE MP', bold)
        worksheet.write('M1', 'Dynamic MP [J/min]', bold)
        worksheet.write('N1', 'SD Dynamic MP', bold)
        worksheet.write('O1', 'SE Dynamic MP', bold)
        worksheet.write('P1', 'Airway HA [J/min]', bold)
        worksheet.write('Q1', 'SD Airway HA', bold)
        worksheet.write('R1', 'SE Airway HA', bold)
        worksheet.write('S1', 'Esophageal PTP [cmH2O.s/min]', bold)
        worksheet.write('T1', 'SD Esophageal PTP', bold)
        worksheet.write('U1', 'SE Esophageal PTP', bold)
        worksheet.write('V1', 'Esophageal WOB [J/min]', bold)
        worksheet.write('W1', 'SD Esophageal WOB', bold)
        worksheet.write('X1', 'SE Esophageal WOB', bold)
        worksheet.write('Y1', 'Transpulmonary PTP [cmH2O.s/min]', bold)
        worksheet.write('Z1', 'SD Transpulmonary PTP', bold)
        worksheet.write('AA1', 'SE Transpulmonary PTP', bold)
        worksheet.write('AB1', 'Transpulmonary WOB [J/min]', bold)
        worksheet.write('AC1', 'SD Transpulmonary WOB', bold)
        worksheet.write('AD1', 'SE Transpulmonary WOB', bold)
        worksheet.write('AE1', 'Transpulmonary HA [J/min]', bold)
        worksheet.write('AF1', 'SD Transpulmonary HA', bold)
        worksheet.write('AG1', 'SE Transpulmonary HA', bold)
        worksheet.write('AH1', 'Transpulmonary Peak Pressure [cmH2O]', bold)
        worksheet.write('AI1', 'SD Transpulmonary Peak Pressure', bold)
        worksheet.write('AJ1', 'SE Transpulmonary Peak Pressure', bold)
        worksheet.write('AK1', 'Transpulmonary Pressure Swing [cmH2O]', bold)
        worksheet.write('AL1', 'SD Transpulmonary Pressure Swing', bold)
        worksheet.write('AM1', 'SE Transpulmonary Pressure Swing', bold)

        # write data to worksheet
        columns_ = (
            input_filename,
            patient_id,
            ventilation_mode_txt,
            insp_hold_txt,
            pressure_type_txt,
            segment_time,
            rr_,
            mean_tidal_volume,
            mean_peep,
            mech_power,
            standard_deviations[0],
            standard_errors[0],
            dyn_mech_power,
            standard_deviations[1],
            standard_errors[1],
            aw_loop_power,
            standard_deviations[2],
            standard_errors[2],
            ptp_es_mean,
            standard_deviations[6],
            standard_errors[6],
            wob_es_mean,
            standard_deviations[5],
            standard_errors[5],
            ptp_tp_mean,
            standard_deviations[7],
            standard_errors[7],
            wob_tp_mean,
            standard_deviations[4],
            standard_errors[4],
            tp_loop_power,
            standard_deviations[3],
            standard_errors[3],
            tp_peak_mean,
            standard_deviations[8],
            standard_errors[8],
            tp_swing_mean,
            standard_deviations[9],
            standard_errors[9],
        )

        # define start row and column
        row = 1
        col = 0

        # iterate over the data and write it out column by column
        for item in columns_:
            worksheet.write(row, col, item)
            col += 1

        workbook.close()

        output_file = 'new file created'
        output_filename = new_output_name + '.xlsx'

    elif output_option == OUTPUT_FILE.EXISTING_FILE:
        output_file = output_xlsx_file[0]

        list_output = [input_filename,
            patient_id,
            ventilation_mode_txt,
            insp_hold_txt,
            pressure_type_txt,
            segment_time,
            rr_,
            mean_tidal_volume,
            mean_peep,
            mech_power,
            standard_deviations[0],
            standard_errors[0],
            dyn_mech_power,
            standard_deviations[1],
            standard_errors[1],
            aw_loop_power,
            standard_deviations[2],
            standard_errors[2],
            ptp_es_mean,
            standard_deviations[6],
            standard_errors[6],
            wob_es_mean,
            standard_deviations[5],
            standard_errors[5],
            ptp_tp_mean,
            standard_deviations[7],
            standard_errors[7],
            wob_tp_mean,
            standard_deviations[4],
            standard_errors[4],
            tp_loop_power,
            standard_deviations[3],
            standard_errors[3],
            tp_peak_mean,
            standard_deviations[8],
            standard_errors[8],
            tp_swing_mean,
            standard_deviations[9],
            standard_errors[9]]

        # re-open and append
        file = load_workbook(output_file)
        ws_ = file.active
        ws_.append(list_output)
        file.save(output_file)
        output_filename = PurePath(output_file).stem + PurePath(output_file).suffix

    return output_file, output_filename
